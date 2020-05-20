#!/user/bin/env python
# -*- coding: utf-8 -*-
# __author__ = yangyd 
# Create time: 2019/7/8 0008 14:25


from openpyxl import load_workbook
from common_code.operate_config import DoConfig
from common_code.project_path import CONFIG_FILE_PATH

conf = DoConfig(CONFIG_FILE_PATH)
actual_col = conf.read_int('test_result', 'actual_col')
test_col = conf.read_int('test_result', 'test_col')


class DoExcel:

    def __init__(self, file_path, sheet_name=None):
        """
        打开文件，并获取工作薄
        :param file_path: 文件路径
        :param sheet_name: sheet名
        """
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_excel(self, start_row=1, start_col=1, end_row=None, end_col=None):
        """
        读取指定行列的所有数据
        :param start_row: 起始行，默认为1
        :param start_col: 起始列，默认为1
        :param end_row: 行数最大值
        :param end_col: 行数最小值
        :return: 返回读取的所有数据
        """
        wb = load_workbook(self.file_path)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]

        if end_row is None:
            end_row = ws.max_row
        else:
            end_row = end_row
        if end_col is None:
            end_col = ws.max_column
        else:
            end_col = end_col

        # 获取表头列，并存储在一个列表内,values_only=True 返回的是值，false的话返回的是cell对象的生成器
        head_value = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = head_value[0]
        # 将指定表内的数据取出，并存储在一个嵌套字典的列表中
        excel_data = []
        for item in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col,
                                 max_col=end_col, values_only=True):
            sub_data = dict(zip(header, item))
            excel_data.append(sub_data)

        return excel_data

    def write_result(self, row, actual_result, test_result):
        """
        写回实际结果和测试结论
        :param row: 写入行号
        :param actual_result:函数运行结果
        :param test_result: 测试结论：pass，fail
        :return:
        """
        wb = load_workbook(self.file_path)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]

        ws.cell(row, actual_col).value = actual_result
        ws.cell(row, test_col).value = test_result
        wb.save(self.file_path)
        wb.close()


if __name__ == '__main__':
    from common_code.project_path import CASE_FILE_PATH
    do_excel = DoExcel(CASE_FILE_PATH, 'login')
    print(do_excel.read_excel(2))
